import marimo

__generated_with = "0.17.7"
app = marimo.App(width="columns")


@app.cell(column=0)
def _():
    import marimo as mo
    import numpy as np
    import xarray as xr
    from pathlib import Path
    from tqdm import tqdm
    import pandas as pd
    import openpyxl
    import polar as pl
    from typing import List, Dict
    import re
    import pyarrow
    import seaborn as sns
    import matplotlib.pyplot as plt
    return List, Path, np, pd, plt, re, sns, tqdm


@app.cell
def _(Path):
    features_dir = Path("/users/thomasbush/Downloads/shared WithTWB/")
    saving_dir = Path("/Users/thomasbush/Downloads/features_plots/")
    saving_dir.mkdir(parents=True, exist_ok=True)
    features_paths = [pathname for pathname in features_dir.iterdir()]
    return features_paths, saving_dir


@app.cell
def _(List, re):
    def find_sessions(pathnames:List=None, condition:str="cricket", mouse:int=1):
        """Filter paths to match desired session conditions"""
        assert mouse > 0, "Mouse starts at 1"
        pattern = rf"m00{mouse}_s[0-9]*_{condition}.xlsx$"
        filtered_paths = [path for path in pathnames if re.match(pattern, path.name)]
        return filtered_paths
    return (find_sessions,)


@app.cell
def _(List, plt, session_m2_cricket_dict):
    feature = "head_speed"
    sessions = list(session_m2_cricket_dict.items())
    n_sessions = len(sessions)

    def plot_time_series(sessions:List, feature:str, n_sessions:int):
        # Create grid: 2 rows, 3 columns (adjust as needed)
        n_rows, n_cols = 2, 3
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 6), sharex=True)
        axes = axes.flatten()  # flatten to iterate easily

        for ax, (sess_name, sess_data) in zip(axes, sessions):
            sess_data[feature].rolling(5).mean().plot(ax=ax)
            # ax.violin_plot()
            ax.set_ylim(0, 400)
            ax.set_title(sess_name)
            ax.set_ylabel(feature)

        # Hide any empty subplots (if fewer sessions than grid cells)
        for ax in axes[n_sessions:]:
            ax.set_visible(False)

        axes[-1].set_xlabel("Frame")
        plt.suptitle(f"{feature} across sessions", y=1.02)
        plt.tight_layout()
        plt.show()
    return (sessions,)


@app.cell
def _(re):
    def parse_filename(filename: str):
        pattern = r"m0*(\d+)_s\d+_(cricket|object)\.xlsx"
        m = re.match(pattern, filename)
        if not m:
            raise ValueError("Filename does not match expected pattern.")
        mouse, condition = m.groups()
        return int(mouse), condition
    return (parse_filename,)


@app.cell
def _(List, np, parse_filename, plt, re):
    def session_index(name: str) -> int:
        """
        Extracts the integer after '_s' in filenames like 'm003_s007_cricket.xlsx'.
        Returns it as an int (e.g. 7).
        """
        m = re.search(r'_s(\d+)', name)
        if m is None:
            # fallback if pattern not found; put these at the end
            return 999999
        return int(m.group(1))

    def plot_violin_sessions(sessions:List, feature_name:str, saving_dir=None, show=True):
        # 1. Sort sessions by the sXXX number
        sorted_sessions = sorted(sessions, key=lambda sess: session_index(sess[0]))

        # 2. Build names + data in that order
        names = []
        data = []
        for name, df in sorted_sessions:
            if feature_name not in df.columns:
                continue

            vals = df[feature_name]
            vals = vals.replace([np.inf, -np.inf], np.nan).dropna().to_numpy()

            if len(vals) > 0:
                names.append(name)
                data.append(vals)

        # 3. Plot
        plt.figure()
        plt.violinplot(data, showmeans=True, showextrema=True, showmedians=True)
        plt.ylabel(feature_name)
        plt.title(f"{feature_name} per session")

        plt.xticks(
            ticks=range(1, len(names) + 1),
            labels=names,
            rotation=90
        )

        plt.tight_layout()
        if saving_dir is not None:
            mouse, condition = parse_filename(names[0])
            violin_plots_dir = saving_dir / "violin_plots"
            violin_plots_dir.mkdir(parents=True, exist_ok = True)
            plt.savefig(violin_plots_dir / f"mouse_{mouse}_{condition}_{feature_name}.png")
        if show:
            if feature_name == "head_speed":
                plt.ylim(0, 200)
            plt.show()
    return plot_violin_sessions, session_index


@app.cell
def _(plot_violin_sessions, sessions):
    plot_violin_sessions(sessions, "head_speed")
    return


@app.cell
def _(np, parse_filename, pd, plt, session_index, sns):
    def plot_feature_dev_heatmap(sessions, saving_dir=None, show=True):

        angle_cols = [
            "facing_angle",
            "rel_bearing",
            "rel_angle_target",
            "angle_head_body_axis",
            "angle_head_body_l",
            "angle_head_body_r",
            "angle_head_body_speed",
            "ori_allBody",
            "ori_trunk",
            "ori_head",
            "omega_theta_trunk",
        ]

        non_angle_cols = [
            "dist_head",
            "dist",
            "dist_change",
            "height",
            "height_scaled",
            "head_acc",
            "rigidbody_acc",
            "speed_head_fwd",
            "head_speed",
            "trunk_speed",
            "radial_vel",
            "tangential_vel",
            "nose_tail_distance",
            "forepawL_tail_distance",
            "forepawR_tail_distance",
            "forepaw_target_distance",
            "dist_head_scaled",
            "dist_scaled",
            "dist_change_scaled",
        ]

        exclude = ["timestamp"]

        all_rows = []
        names = []

        for name, df in sessions:
            sid = session_index(name)
            names.append(name)
            if sid is None:
                continue

            row = df.select_dtypes(include=[np.number]).mean()
            row["session"] = f"s{sid:03d}"
            row["session_id"] = sid
            all_rows.append(row)

        features_df = pd.DataFrame(all_rows)
        features_df = features_df.sort_values("session_id").drop(columns="session_id")

        num_df = features_df.drop(columns=exclude).set_index("session")

        angle_df = num_df[[c for c in angle_cols if c in num_df.columns]]
        non_angle_df = num_df[[c for c in non_angle_cols if c in num_df.columns]]

        # Z-score non-angle features
        Z = (non_angle_df - non_angle_df.mean()) / non_angle_df.std()
        Z_filtered = Z.loc[:, Z.var() > 1e-3]

        # Plot
        fig, axes = plt.subplots(
            2 if not angle_df.empty else 1,
            1,
            figsize=(12, 10 if not angle_df.empty else 6),
            sharex=True
        )

        if isinstance(axes, np.ndarray):
            ax1, ax2 = axes
        else:
            ax1 = axes
            ax2 = None

        # non-angle heatmap
        sns.heatmap(
            Z_filtered.T,
            cmap="magma",
            center=0,
            cbar_kws={'label': 'Z-score'},
            ax=ax1
        )
        ax1.set_title("Non-angle Features (Z-scored)")
        ax1.set_ylabel("Feature")

        # angle heatmap
        if ax2 is not None:
            sns.heatmap(
                angle_df.T,
                cmap="coolwarm",
                center=0,
                cbar_kws={'label': 'Angle (deg)'},
                ax=ax2,
            )
            ax2.set_title("Angle Features")
            ax2.set_xlabel("Session")
            ax2.set_ylabel("Angle Feature")

        plt.tight_layout()

        if saving_dir is not None:
            mouse, condition = parse_filename(names[0])
            outdir = saving_dir / "heatmaps"
            outdir.mkdir(parents=True, exist_ok=True)
            plt.savefig(outdir / f"mouse_{mouse}_{condition}.png")

        if show:
            plt.show()
    return (plot_feature_dev_heatmap,)


@app.cell
def _(plot_feature_dev_heatmap, sessions):
    plot_feature_dev_heatmap(sessions)
    return


@app.cell
def _(Path):
    # plot occupancy
    from movement.plots import plot_occupancy
    import pickle

    coordinates_path = Path("/Users/thomasbush/Downloads/multicam_video_2025-05-07T12_16_20_cropped-v2_20250701121021/multicam_video_2025-05-07T12_16_20_centralDLC_HrnetW48_mouse-bottomJul1shuffle1_snapshot_189_full.pickle")

    with open(coordinates_path, "rb") as f:
        data = pickle.load(f)
    print(len(data))

    return data, plot_occupancy


@app.cell
def _(data, np, tqdm):
    coordinates = []
    confidences = []

    for frame, values in tqdm(data.items()):
        if frame == "metadata":
            continue
        coordinates.append(values["coordinates"])
        confidences.append(np.asarray(values["confidence"]))

    coordinates = np.vstack(coordinates)
    confidences = np.stack(confidences, axis=1)
    return confidences, coordinates


@app.cell
def _():
    bodyparts = [
      "belly_caudal",
      "belly_rostral",
      "ear_lf",
      "ear_rt",
      "forepaw_lf",
      "forepaw_rt",
      "hindpaw_lf",
      "hindpaw_rt",
      "nose",
      "tailbase"
    ]
    return (bodyparts,)


@app.cell
def _(bodyparts, confidences, coordinates):
    from movement.io.load_poses import from_numpy

    xr_data = from_numpy(
        coordinates.transpose(0, 3, 1, 2),
        confidence_array=confidences.transpose(1, 0, 2, 3)[..., 0],
        individual_names = "mouse",
        fps=60,
        keypoint_names=bodyparts,source_software="dlc"
    )
    return (xr_data,)


@app.cell
def _(plot_occupancy, xr_data):
    plot_occupancy(xr_data.position, keypoints=["ear_lf", "ear_rt", "nose"], bins=[50, 50], cmin=1, norm="log")
    return


@app.cell
def _(xr_data):
    from movement.plots import plot_centroid_trajectory

    plot_centroid_trajectory(xr_data.position, keypoints=['nose', 'ear_lf', 'ear_rt'])
    return


@app.cell
def _():
    return


@app.cell(column=1)
def _(features_paths, parse_filename):
    # get unique mice and conditions:
    mice = []
    conditions = ["cricket", "object"]
    for pathname in features_paths:
        if pathname.name == "startTimes.xlsx":
            continue
        m, condition = parse_filename(pathname.name)
        mice.append(m)
    mice = set(mice)
    return


@app.cell
def _(
    List,
    Path,
    find_sessions,
    parse_filename,
    pd,
    plot_feature_dev_heatmap,
    plot_violin_sessions,
    tqdm,
):
    def generate_and_save_plots(features_paths:List,feature_list=None, saving_dir:Path=None):
        # extract unique conditions and mice
        mice = []
        conditions = ["cricket", "object"]
        for pathname in features_paths:
            if pathname.name == "startTimes.xlsx":
                continue
            m, condition = parse_filename(pathname.name)
            mice.append(m)
        mice = set(mice)
        # deal with features for the violin plots
        if not feature_list:
            feature_list = ["head_speed", "facing_angle"]

        for m in tqdm(mice):
            for condition in conditions:
                temp_sessions = find_sessions(features_paths, condition, m)
                temp_dict = {path.name: pd.read_excel(path) for path in temp_sessions}
                sessions_list = list(temp_dict.items())
                n_sessions = len(sessions_list)

                for feature_ in feature_list:
                    plot_violin_sessions(sessions_list, feature_, saving_dir, show=False)
                plot_feature_dev_heatmap(sessions_list, saving_dir, show=False)
        print("All Plots generated:")
    return (generate_and_save_plots,)


@app.cell
def _(
    features_paths,
    generate_and_save_plots,
    saving_dir,
    session_m2_cricket_dict,
):
    features_all = session_m2_cricket_dict["m003_s001_cricket.xlsx"].columns.tolist()[1:]
    generate_and_save_plots(features_paths, saving_dir=saving_dir, feature_list=features_all)
    return


@app.cell
def _():
    return


@app.cell(column=2)
def _():
    import statsmodels.formula.api as smf
    return (smf,)


@app.cell
def _(parse_filename, pd, session_index):
    def build_session_level_df(features_paths, feature_list=None):
        rows = []

        for path in features_paths:
            if path.name == "startTimes.xlsx":
                continue

            # parse mouse + condition from filename: m001_s001_cricket.xlsx
            mouse, condition = parse_filename(path.name)  # you already have this
            session_id = session_index(path.name)        # you already have this

            df = pd.read_excel(path)

            # decide which features to keep
            if feature_list is None:
                # default: all numeric columns except timestamp
                numeric_cols = df.select_dtypes(include="number").columns
                feature_cols = [c for c in numeric_cols if c != "timestamp"]
            else:
                feature_cols = feature_list

            # compute session-level means
            feature_means = df[feature_cols].mean()

            row = {
                "mouse": mouse,
                "condition": condition,
                "session": session_id,
            }
            row.update(feature_means.to_dict())
            rows.append(row)

        session_df = pd.DataFrame(rows)

        # make sure types are nice
        session_df["mouse"] = session_df["mouse"].astype(int)
        session_df["condition"] = session_df["condition"].astype("category")
        session_df["session"] = session_df["session"].astype(int)

        # sort for sanity
        session_df = session_df.sort_values(["mouse", "condition", "session"])

        return session_df
    return (build_session_level_df,)


@app.cell
def _(build_session_level_df, features_paths):
    feature_list = ["dist_head", "head_speed", "facing_angle", "speed_head_fwd", "head_acc", "nose_tail_distance"]  # or None for all numeric
    session_df = build_session_level_df(features_paths, feature_list)

    session_df.head()
    return (session_df,)


@app.cell
def _(np, session_df):
    session_df["cos_facing_angle"] = np.cos(session_df["facing_angle"])
    return


@app.cell
def _(session_df, smf):
    model = smf.mixedlm(
        "dist_head ~ session * condition",
        data=session_df,
        groups=session_df["mouse"],
    )
    res = model.fit()
    print(res.summary())
    return


app._unparsable_cell(
    r"""
    !pip install ssm
    """,
    name="_"
)


@app.cell
def _():
    return


if __name__ == "__main__":
    app.run()
